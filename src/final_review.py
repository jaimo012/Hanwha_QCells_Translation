"""
최종검수 모듈 (Final Review Module)

완료 폴더(completed_folder)의 번역 파일들을 최종 검수합니다.
Google Sheets '최종검수' 시트의 파일 목록을 순차적으로 처리하며,
각 파일에 대해 다음 항목을 확인합니다:

1. 원본 파일 존재 여부 (F열)
2. 번역본(" - en") 파일 존재 여부 (G열) + 파일명 기록 (E열)
3. 원본 파일 오픈 가능 여부 (H열)
4. 번역본 파일 오픈 가능 여부 (I열)
5. 번역 완료 여부 - 한글 잔존 확인 (J열)
6. 검수 일시 기록 (K열)
"""

import gc
import os
import time
import traceback
from datetime import datetime

from docx import Document
from pptx import Presentation
import openpyxl
import gspread
from google.oauth2.service_account import Credentials

from .config import (
    PROJECT_ROOT,
    COMPLETED_FOLDER,
    GOOGLE_SHEETS_URL,
    validate_config,
)
from .verify import (
    build_work_file_path,
    scan_korean_in_file,
)


# ==============================================================================
# [상수 정의] 최종검수 시트 설정
# ==============================================================================
FINAL_REVIEW_SHEET_NAME = "최종검수"

# Google Sheets API Rate Limit 대응 설정
SHEETS_API_MIN_DELAY = 0.5       # API 호출 간 최소 대기 시간 (초)
SHEETS_API_RETRY_COUNT = 3       # 재시도 횟수
SHEETS_API_RETRY_DELAY = 5       # 재시도 대기 시간 (초)


class FinalReviewColumns:
    """
    최종검수 시트 컬럼 인덱스 (1-based)

    | A(1) | B(2)   | C(3)   | D(4)     | E(5)     | F(6)     | G(7)     | H(8)     | I(9)     | J(10)    | K(11)      |
    | 연번 | 상위경로 | 세부경로 | 원본파일명 | 번역본파일명 | 원본파일여부 | 번역본파일여부 | 원본오픈상태 | 번역본오픈상태 | 번역완료여부 | 최종검수일시 |
    """
    ROW_NUM = 1             # A: 연번
    UPPER_PATH = 2          # B: 상위경로
    SUB_PATH = 3            # C: 세부경로
    ORIGINAL_FILE = 4       # D: 원본 파일명
    TRANSLATED_FILE = 5     # E: 번역본 파일명
    ORIGINAL_EXISTS = 6     # F: 원본 파일여부
    TRANSLATED_EXISTS = 7   # G: 번역본 파일여부
    ORIGINAL_OPENS = 8      # H: 원본 오픈상태
    TRANSLATED_OPENS = 9    # I: 번역본 오픈상태
    TRANSLATION_DONE = 10   # J: 번역완료여부
    REVIEW_DATETIME = 11    # K: 최종검수일시


# ==============================================================================
# [Google Sheets 연결]
# ==============================================================================

def connect_to_review_sheet():
    """
    Google Sheets의 '최종검수' 시트에 연결합니다.

    동일 스프레드시트 내 다른 탭(시트)에 접근하기 위해
    기존 SheetsManager와 별도로 gspread를 직접 사용합니다.

    Returns:
        gspread.Worksheet: 최종검수 시트 객체

    Raises:
        Exception: 연결 실패 시
    """
    credentials_path = os.path.join(PROJECT_ROOT, "credentials.json")

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    creds = Credentials.from_service_account_file(credentials_path, scopes=scopes)
    client = gspread.authorize(creds)

    spreadsheet = client.open_by_url(GOOGLE_SHEETS_URL)
    sheet = spreadsheet.worksheet(FINAL_REVIEW_SHEET_NAME)

    print(f"✅ Google Sheets 연결 성공 (시트: {FINAL_REVIEW_SHEET_NAME})")
    return sheet


def api_call_with_retry(func, *args, **kwargs):
    """
    Google Sheets API Rate Limit(429) 대응 재시도 래퍼 함수

    Args:
        func: 실행할 함수
        *args, **kwargs: 함수 인자

    Returns:
        함수 실행 결과

    Raises:
        Exception: 최대 재시도 후에도 실패 시
    """
    last_exception = None

    for attempt in range(SHEETS_API_RETRY_COUNT):
        try:
            time.sleep(SHEETS_API_MIN_DELAY)
            return func(*args, **kwargs)
        except Exception as e:
            last_exception = e
            error_str = str(e)

            if '429' in error_str or 'Quota exceeded' in error_str:
                wait_time = SHEETS_API_RETRY_DELAY * (attempt + 1)
                print(f"   ⏳ API 한도 초과, {wait_time}초 대기 후 재시도 ({attempt + 1}/{SHEETS_API_RETRY_COUNT})...")
                time.sleep(wait_time)
            else:
                raise

    raise last_exception


# ==============================================================================
# [파일 검색 함수]
# ==============================================================================

def find_original_file(upper_path, sub_path, file_name):
    """
    완료 폴더(completed_folder)에서 원본 파일을 찾습니다.

    경로 구성: completed_folder / 상위경로 / 세부경로 / 원본파일명

    Args:
        upper_path (str): 상위 경로 (예: "MES")
        sub_path (str): 세부 경로 (예: "30. 개발단계/10. 기능설계서")
        file_name (str): 원본 파일명 (확장자 포함)

    Returns:
        tuple: (파일 존재 여부 bool, 파일 전체 경로 str)
    """
    file_path = os.path.join(COMPLETED_FOLDER, upper_path, sub_path, file_name)
    exists = os.path.exists(file_path)
    return exists, file_path


def find_translated_file(upper_path, sub_path, file_name):
    """
    번역 작업본 파일(" - en")을 찾습니다.

    여러 가능한 파일명 변형을 순서대로 시도합니다:
    1. build_work_file_path 표준 경로 (doc→docx, 확장자 소문자 변환)
    2. 원본 확장자 그대로 사용한 경우 (예: .PPTX 그대로)
    3. 디렉토리 내 패턴 매칭 검색 (대소문자 무시)

    Args:
        upper_path (str): 상위 경로
        sub_path (str): 세부 경로
        file_name (str): 원본 파일명

    Returns:
        tuple: (존재 여부 bool, 파일 전체 경로 str or None, 파일명 str or None)
    """
    # --- 시도 1: 표준 경로 (doc→docx, 확장자 소문자) ---
    standard_path = build_work_file_path(upper_path, sub_path, file_name)
    if os.path.exists(standard_path):
        return True, standard_path, os.path.basename(standard_path)

    # --- 시도 2: 원본 확장자 그대로 사용한 경우 ---
    name, ext = os.path.splitext(file_name)

    # 대문자 확장자를 그대로 사용했을 수 있음 (예: .PPTX)
    if ext != ext.lower():
        original_ext_path = os.path.join(
            COMPLETED_FOLDER, upper_path, sub_path, f"{name} - en{ext}"
        )
        if os.path.exists(original_ext_path):
            return True, original_ext_path, os.path.basename(original_ext_path)

    # --- 시도 3: 디렉토리에서 패턴 매칭 검색 ---
    target_dir = os.path.join(COMPLETED_FOLDER, upper_path, sub_path)

    if not os.path.exists(target_dir):
        return False, None, None

    name_lower = name.lower()

    try:
        for entry in os.listdir(target_dir):
            entry_name_part, _ = os.path.splitext(entry)
            entry_name_lower = entry_name_part.lower()

            # 원본 파일명 + " - en" 패턴 매칭 (대소문자 무시)
            if entry_name_lower == f"{name_lower} - en":
                found_path = os.path.join(target_dir, entry)
                return True, found_path, entry
    except OSError as e:
        print(f"      ⚠️ 디렉토리 검색 오류: {e}")

    return False, None, None


# ==============================================================================
# [파일 오픈 검사]
# ==============================================================================

def try_open_file(file_path):
    """
    파일을 실제로 열어 오픈 가능 여부를 확인합니다.

    파일 형식별 검사 방법:
    - .docx: python-docx로 Document를 열고 paragraphs 접근
    - .pptx: python-pptx로 Presentation을 열고 slides 접근
    - .xlsx: openpyxl로 load_workbook 후 sheetnames 접근
    - .doc:  구버전 Word 형식은 파일 크기로 기본 체크
    - 기타:  파일 크기로 기본 체크

    ⚠️ 중요: 열었던 파일 객체는 반드시 닫거나 삭제하여
    파일 핸들 누수와 메모리 누적을 방지합니다.

    Args:
        file_path (str): 검사할 파일 경로

    Returns:
        bool: 정상적으로 열리면 True, 아니면 False
    """
    if not file_path or not os.path.exists(file_path):
        return False

    try:
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            return False

        ext = os.path.splitext(file_path)[1].lower()

        if ext == '.docx':
            doc = Document(file_path)
            _ = doc.paragraphs
            del doc  # 메모리에서 명시적 해제
            return True

        if ext == '.pptx':
            prs = Presentation(file_path)
            _ = prs.slides
            del prs  # 메모리에서 명시적 해제
            return True

        if ext == '.xlsx':
            wb = openpyxl.load_workbook(file_path, read_only=True)
            try:
                _ = wb.sheetnames
                return True
            finally:
                # read_only 모드는 파일 핸들을 계속 잡고 있으므로
                # 성공/실패와 무관하게 반드시 닫아야 함
                wb.close()

        if ext == '.doc':
            # .doc (구버전 Word)는 python-docx로 열 수 없음
            # 파일 크기 확인으로 기본 무결성 체크
            return file_size > 0

        # 기타 형식: 파일 크기로 기본 체크
        return file_size > 0

    except Exception as e:
        print(f"      ⚠️ 파일 열기 실패: {e}")
        return False


# ==============================================================================
# [번역 완료 여부 확인]
# ==============================================================================

def check_translation_complete(file_path):
    """
    번역 파일에 한글이 남아있는지 확인합니다.

    verify.py의 scan_korean_in_file()을 활용하여
    docx/pptx/xlsx 파일 내부의 한글 잔존 여부를 검사합니다.

    Args:
        file_path (str): 번역본 파일 경로

    Returns:
        bool: 한글이 없으면 True (번역 완료), 한글이 있으면 False
    """
    if not file_path or not os.path.exists(file_path):
        return False

    try:
        has_korean_text, korean_count = scan_korean_in_file(file_path)

        if has_korean_text:
            print(f"      📝 한글 잔존: {korean_count}개 항목")
            return False

        return True

    except Exception as e:
        print(f"      ⚠️ 한글 검사 오류: {e}")
        return False


# ==============================================================================
# [시트 업데이트]
# ==============================================================================

def update_row_result(sheet, row_index, results):
    """
    한 행의 검수 결과를 시트에 업데이트합니다.

    E열(번역본 파일명) ~ K열(최종검수일시)을 한 번의 API 호출로
    효율적으로 업데이트합니다.

    Args:
        sheet (gspread.Worksheet): 시트 객체
        row_index (int): 행 번호 (1-based)
        results (dict): 검수 결과 딕셔너리
    """
    cell_range = f"E{row_index}:K{row_index}"

    values = [[
        results.get('translated_file_name', ''),        # E: 번역본 파일명
        str(results['original_exists']),                 # F: 원본 파일여부
        str(results['translated_exists']),               # G: 번역본 파일여부
        str(results['original_opens']),                  # H: 원본 오픈상태
        str(results['translated_opens']),                # I: 번역본 오픈상태
        str(results['translation_done']),                # J: 번역완료여부
        results['review_datetime'],                      # K: 최종검수일시
    ]]

    api_call_with_retry(sheet.update, values, range_name=cell_range)


# ==============================================================================
# [단일 행 검수 처리]
# ==============================================================================

def review_single_row(sheet, row_index, upper_path, sub_path, file_name):
    """
    단일 행의 최종검수를 수행합니다 (1~7번 단계 전체).

    Args:
        sheet (gspread.Worksheet): 시트 객체
        row_index (int): 행 번호 (1-based)
        upper_path (str): 상위 경로
        sub_path (str): 세부 경로
        file_name (str): 원본 파일명

    Returns:
        dict: 검수 결과 딕셔너리
    """
    results = {
        'translated_file_name': '',
        'original_exists': False,
        'translated_exists': False,
        'original_opens': False,
        'translated_opens': False,
        'translation_done': False,
        'review_datetime': '',
    }

    # ── Step 1~2: 원본 파일 찾기 → F열 ──
    original_exists, original_path = find_original_file(upper_path, sub_path, file_name)
    results['original_exists'] = original_exists

    if original_exists:
        print(f"      ✅ F: 원본 파일 존재")
    else:
        print(f"      ❌ F: 원본 파일 없음")

    # ── Step 3: 번역 작업본(" - en") 찾기 → E열, G열 ──
    translated_exists, translated_path, translated_name = find_translated_file(
        upper_path, sub_path, file_name
    )
    results['translated_exists'] = translated_exists

    if translated_exists:
        results['translated_file_name'] = translated_name
        print(f"      ✅ G: 번역본 발견 → {translated_name}")
    else:
        print(f"      ❌ G: 번역본 없음")

    # ── Step 4: 원본 파일 오픈 확인 → H열 ──
    if original_exists:
        original_opens = try_open_file(original_path)
        results['original_opens'] = original_opens
        icon = "✅" if original_opens else "❌"
        print(f"      {icon} H: 원본 오픈 {'성공' if original_opens else '실패'}")
    else:
        print(f"      ⏭️  H: 원본 파일 없어 오픈 불가")

    # ── Step 5: 번역본 파일 오픈 확인 → I열 ──
    if translated_exists:
        translated_opens = try_open_file(translated_path)
        results['translated_opens'] = translated_opens
        icon = "✅" if translated_opens else "❌"
        print(f"      {icon} I: 번역본 오픈 {'성공' if translated_opens else '실패'}")
    else:
        print(f"      ⏭️  I: 번역본 없어 오픈 불가")

    # ── Step 6: 번역 완료 여부 (한글 잔존 확인) → J열 ──
    if translated_exists and results['translated_opens']:
        translation_done = check_translation_complete(translated_path)
        results['translation_done'] = translation_done
        icon = "✅" if translation_done else "⚠️"
        status_text = "번역 완료 (한글 없음)" if translation_done else "한글 잔존 (재번역 필요)"
        print(f"      {icon} J: {status_text}")
    elif not translated_exists:
        print(f"      ⏭️  J: 번역본 없어 확인 불가")
    else:
        print(f"      ⏭️  J: 번역본 열리지 않아 확인 불가")

    # ── Step 7: 검수일시 기록 → K열 ──
    results['review_datetime'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── 시트에 결과 기록 (E~K열, 단일 API 호출) ──
    update_row_result(sheet, row_index, results)
    print(f"      📝 K: 시트 기록 완료 ({results['review_datetime']})")

    # ── 리소스 정리: 이 파일에서 열었던 모든 객체를 메모리에서 해제 ──
    # python-docx, python-pptx, openpyxl 등이 남긴 객체를 정리
    gc.collect()

    return results


# ==============================================================================
# [검수 대상 행 필터링]
# ==============================================================================

def get_pending_rows(all_values):
    """
    시트 데이터에서 검수가 필요한 행을 필터링합니다.

    K열(최종검수일시)이 비어있는 행만 검수 대상으로 선정합니다.
    이미 검수된 행(K열에 값이 있는 행)은 건너뜁니다.

    Args:
        all_values (list): 시트의 전체 데이터 (헤더 포함)

    Returns:
        tuple: (검수 대상 행 리스트, 이미 완료된 행 수)
    """
    col = FinalReviewColumns
    pending_rows = []
    already_done_count = 0

    for idx, row in enumerate(all_values[1:], start=2):
        # D열(원본 파일명)이 없으면 빈 행으로 간주
        file_name = row[col.ORIGINAL_FILE - 1] if len(row) >= col.ORIGINAL_FILE else ''
        if not file_name.strip():
            continue

        # K열(최종검수일시) 확인 - 값이 있으면 이미 검수 완료
        review_datetime = row[col.REVIEW_DATETIME - 1] if len(row) >= col.REVIEW_DATETIME else ''
        if review_datetime.strip():
            already_done_count += 1
            continue

        upper_path = row[col.UPPER_PATH - 1] if len(row) >= col.UPPER_PATH else ''
        sub_path = row[col.SUB_PATH - 1] if len(row) >= col.SUB_PATH else ''

        pending_rows.append({
            'row_index': idx,
            'upper_path': upper_path.strip(),
            'sub_path': sub_path.strip(),
            'file_name': file_name.strip(),
        })

    return pending_rows, already_done_count


# ==============================================================================
# [메인 함수]
# ==============================================================================

def main():
    """
    최종검수 프로세스 메인 함수

    Google Sheets '최종검수' 시트의 파일 목록을 순차 처리하며,
    각 파일의 존재 여부, 오픈 상태, 번역 완료 여부를 확인하고
    결과를 시트에 기록합니다.
    """
    print("=" * 60)
    print("📋 한화큐셀 번역 프로젝트 - 최종검수 프로세스")
    print("   완료 파일의 존재/오픈/번역완료 여부를 검사합니다")
    print("=" * 60)

    # ── 1. 설정 검증 ──
    is_valid, message = validate_config()
    if not is_valid:
        print(f"\n❌ 설정 오류: {message}")
        return

    print("\n✅ 설정 검증 완료")
    print(f"   📁 완료 폴더: {COMPLETED_FOLDER}")

    # ── 2. Google Sheets 연결 ──
    try:
        sheet = connect_to_review_sheet()
    except Exception as e:
        print(f"\n❌ Google Sheets 연결 실패: {e}")
        return

    # ── 3. 시트 데이터 읽기 ──
    print("\n📊 시트 데이터 읽는 중...")
    try:
        all_values = api_call_with_retry(sheet.get_all_values)
    except Exception as e:
        print(f"❌ 시트 데이터 읽기 실패: {e}")
        return

    total_data_rows = len(all_values) - 1  # 헤더 제외
    if total_data_rows <= 0:
        print("⚠️ 시트에 데이터가 없습니다 (헤더만 존재)")
        return

    print(f"   📋 전체 {total_data_rows}개 행 발견")

    # ── 4. 검수 대상 필터링 ──
    pending_rows, already_done_count = get_pending_rows(all_values)

    if not pending_rows:
        print("\n" + "=" * 60)
        print("✅ 모든 행의 검수가 완료되었습니다!")
        print(f"   (전체 {total_data_rows}개 중 {already_done_count}개 검수 완료)")
        print("=" * 60)
        return

    print(f"   ✅ 검수 완료: {already_done_count}개")
    print(f"   🔍 검수 대상: {len(pending_rows)}개")

    # ── 5. 검수 루프 ──
    success_count = 0
    error_count = 0

    # 결과 통계 (True/False 각각의 개수)
    stats = {
        'original_exists': {True: 0, False: 0},
        'translated_exists': {True: 0, False: 0},
        'original_opens': {True: 0, False: 0},
        'translated_opens': {True: 0, False: 0},
        'translation_done': {True: 0, False: 0},
    }

    print(f"\n🚀 최종검수 시작... ({len(pending_rows)}개 파일)")
    print("   (Ctrl+C로 중단할 수 있습니다)\n")

    for i, row_info in enumerate(pending_rows, start=1):
        try:
            row_index = row_info['row_index']
            upper_path = row_info['upper_path']
            sub_path = row_info['sub_path']
            file_name = row_info['file_name']

            progress = f"[{i}/{len(pending_rows)}]"
            print(f"{'─' * 60}")
            print(f"   {progress} {file_name}")
            print(f"   경로: {upper_path}/{sub_path}")

            # 검수 실행
            results = review_single_row(
                sheet, row_index, upper_path, sub_path, file_name
            )

            # 통계 누적
            for key in stats:
                value = results[key]
                stats[key][value] += 1

            success_count += 1

        except KeyboardInterrupt:
            print("\n\n⚠️ 사용자에 의해 중단되었습니다.")
            break

        except Exception as e:
            print(f"      ❌ 검수 오류: {e}")
            traceback.print_exc()
            error_count += 1

    # ── 6. 최종 결과 출력 ──
    print("\n" + "=" * 60)
    print("📊 최종검수 완료 요약")
    print("=" * 60)
    print(f"   처리 완료: {success_count}개 / 오류: {error_count}개")
    print()

    stat_labels = {
        'original_exists': '📁 원본 파일 존재 (F열)',
        'translated_exists': '📁 번역본 파일 존재 (G열)',
        'original_opens': '📂 원본 오픈 상태 (H열)',
        'translated_opens': '📂 번역본 오픈 상태 (I열)',
        'translation_done': '✅ 번역 완료 여부 (J열)',
    }

    for key, label in stat_labels.items():
        true_count = stats[key][True]
        false_count = stats[key][False]
        print(f"   {label}:  True {true_count}개 / False {false_count}개")

    print("=" * 60)


if __name__ == "__main__":
    main()
